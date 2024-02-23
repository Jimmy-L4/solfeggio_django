import openpyxl
import matplotlib.pyplot as plt
import numpy as np
plt.rcParams['font.sans-serif'] = ['SimHei']  # 设置中文字体为黑体

# 打开 Excel 文件
workbook = openpyxl.load_workbook('2023-汇总.xlsx')
sheet = workbook.active

# 提取数据
average_online_time = []
total_score = []
for row in range(2, sheet.max_row + 1):  # 从第二行开始，因为第一行是表头
    average_online_time.append(sheet.cell(row, 4).value)  # 平均在线时长列
    total_score.append(sheet.cell(row, 5).value)  # 总成绩列，假设为第6列

# 绘制散点图
plt.figure(figsize=(8, 6))
plt.scatter(average_online_time, total_score, alpha=0.5)
plt.title('学生在线时长及成绩关系散点图')
plt.xlabel('在线时长')
plt.ylabel('总成绩')
plt.grid(True)

# 对散点进行拟合并绘制拟合曲线
coefficients = np.polyfit(average_online_time, total_score, 1)  # 使用一次多项式进行拟合
polynomial = np.poly1d(coefficients)
x_values = np.linspace(min(average_online_time), max(average_online_time), 100)
y_values = polynomial(x_values)
plt.plot(x_values, y_values, color='red')

plt.show()
